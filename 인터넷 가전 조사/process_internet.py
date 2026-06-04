import openpyxl
import re
import shutil
import os


# ─────────────────────────────────────────────
# 기본 숫자 추출
# ─────────────────────────────────────────────
def extract_amounts(text):
    """텍스트에서 금액 목록 반환 (만원/쉼표/숫자 형태 처리)"""
    if not text:
        return []
    amounts = []
    for m in re.finditer(r'(\d+(?:\.\d+)?)\s*만\s*원?', text):
        amounts.append(int(float(m.group(1)) * 10000))
    for m in re.finditer(r'(\d{1,3}(?:,\d{3})+)\s*원?', text):
        v = int(m.group(1).replace(',', ''))
        if v not in amounts:
            amounts.append(v)
    for m in re.finditer(r'(?<!\d)(\d{5,6})(?!\d)', text):
        v = int(m.group(1))
        if v not in amounts:
            amounts.append(v)
    return amounts


def amount_near_keyword(text, keyword_regex, look_before=25, look_after=40):
    """
    키워드 위치 주변에서 '가장 가까운' 금액 반환.
    여러 키워드 매칭 시 각각의 근처 금액 목록 반환.
    """
    results = []
    text_lo = text.lower()
    for m in re.finditer(keyword_regex, text_lo):
        kw_s = m.start()
        kw_e = m.end()
        win_s = max(0, kw_s - look_before)
        win_e = min(len(text), kw_e + look_after)
        snippet = text[win_s:win_e]
        kw_pos = kw_s - win_s  # keyword start in snippet

        best, best_d = None, float('inf')
        for mm in re.finditer(r'(\d+(?:\.\d+)?)\s*만\s*원?', snippet):
            v = int(float(mm.group(1)) * 10000)
            d = abs(mm.start() - kw_pos)
            if d < best_d:
                best_d, best = d, v
        for mm in re.finditer(r'(\d{1,3}(?:,\d{3})+)\s*원?', snippet):
            v = int(mm.group(1).replace(',', ''))
            d = abs(mm.start() - kw_pos)
            if d < best_d:
                best_d, best = d, v
        # 쉼표 없는 5-6자리 숫자 (예: 30000, 20000) — \b는 한국어와 경계 문제 있으므로 (?<!\d) 사용
        for mm in re.finditer(r'(?<!\d)(\d{5,6})(?!\d)', snippet):
            v = int(mm.group(1))
            d = abs(mm.start() - kw_pos)
            if d < best_d:
                best_d, best = d, v
        if best is not None:
            results.append(best)
    return results


def parse_money_cell(cell):
    """현금/상품권 셀 파싱 — 숫자 또는 텍스트 혼합"""
    if cell is None:
        return 0
    s = str(cell).strip()
    if s.upper() in ('NONE', 'X', '없음', '없음.', '없음..', '없음 ', 'NAN', '-', ''):
        return 0
    try:
        return int(float(s))
    except (ValueError, TypeError):
        pass
    m = re.search(r'=\s*(\d+)', s)
    if m:
        v = int(m.group(1))
        return v * 10000 if v < 1000 else v
    amounts = extract_amounts(s)
    return max(amounts) if amounts else 0


# ─────────────────────────────────────────────
# 추가현금 — col11 전용
# ─────────────────────────────────────────────
_SKIP_INFO = re.compile(
    r'공유기|와이파이|wifi|월\s*요금|약정|월정액|채널|셋톱|결합|통신사|알뜰폰|첫\s*달|부과|vat|'
    r'3년|2년|1년|공유기|무선|임대|할증|할인요금|월\s*\d{2,3},\d{3}',
    re.IGNORECASE
)

def parse_additional_cash(col11):
    """
    그 외 추가 혜택(col11)에서만 추가현금 추출.
    • '추가 사은품 X' → 추가현금
    • '당일가입/당일접수 X' → 추가현금 (후기와 같은 줄이어도 먼저 처리)
    • '설치비 X 지원/무상' → 설치비를 업체가 커버 → 추가현금
    • '추가지원금 X', 'X만원 추가', 'X원 추가' → 추가현금
    • 일반 '사은품 X' (= 기본 혜택 요약) → 제외
    • 후기/리뷰/카페/블로그/추천만 있는 줄 → 제외
    """
    if not col11:
        return 0
    extra = str(col11)
    total = 0

    # ── 1. 당일 가입/접수 금액 → 바로 뒤 10자 이내 금액만 (같은 줄 리뷰 금액 오염 방지) ──
    for m in re.finditer(r'당일\s*(가입|접수).{0,10}?(\d+(?:\.\d+)?)\s*만\s*원?', extra):
        total += int(float(m.group(2)) * 10000)
    for m in re.finditer(r'당일\s*(가입|접수).{0,10}?(\d{1,3}(?:,\d{3})+)\s*원?', extra):
        total += int(m.group(2).replace(',', ''))

    # ── 2. 줄 단위 처리 ──
    for line in extra.split('\n'):
        line = line.strip()
        if not line:
            continue
        lo = line.lower()

        # 당일 관련은 위에서 처리했으므로 스킵
        if re.search(r'당일\s*(가입|접수)', lo):
            continue

        # 후기·리뷰·추천 줄 → 리뷰추천 영역, 여기서 제외
        if re.search(r'후기|리뷰|카페|블로그|친구\s*추천|지인\s*추천', lo):
            continue

        # 일반 정보성 줄 제외
        if _SKIP_INFO.search(lo):
            continue

        # 설치비 — 지원이면 추가현금, 아니면 무시
        if re.search(r'설치비|출장비|기사출동비', lo):
            if re.search(r'지원|무상|무료', lo):
                # look_before=15로 "X원 설치비 지원" 형태도 커버
                v_list = amount_near_keyword(line, r'설치비|출장비|기사출동비', 15, 25)
                if v_list:
                    total += max(v_list)
            continue

        # 이벤트 사은품 → 리뷰추천 영역, 제외
        if re.search(r'이벤트\s*사은품', lo):
            continue

        # 추가 사은품 X → 추가현금 ✓
        if re.search(r'추가\s*사은품', lo):
            v_list = amount_near_keyword(line, r'사은품', 5, 30)
            if v_list:
                total += max(v_list)
            continue

        # 일반 사은품 X (추가 없이) → 기본 혜택 요약이므로 제외
        if re.search(r'사은품', lo) and not re.search(r'추가', lo):
            continue

        # "추가지원금 X" / "현금 추가 X" 패턴
        if re.search(r'추가지원금|추가\s*현금|현금\s*추가', lo):
            v_list = amount_near_keyword(line, r'추가|현금', 5, 30)
            if v_list:
                total += max(v_list)
            continue

        # "X만원추가" / "X원 추가" 형태
        for m in re.finditer(r'(\d+(?:\.\d+)?)\s*만\s*원?\s*추가', line):
            total += int(float(m.group(1)) * 10000)
        for m in re.finditer(r'(\d{1,3}(?:,\d{3})+)\s*원\s*추가', line):
            total += int(m.group(1).replace(',', ''))

    return total


# ─────────────────────────────────────────────
# 리뷰+추천 보너스
# ─────────────────────────────────────────────
def parse_review_referral(col11, col12):
    """
    col11 우선, 없으면 col12 fallback.
    • 후기/리뷰/카페/블로그 → 리뷰
    • 친구추천/지인추천 → 추천
    • 카페(X)/블로그(Y) '/' → max(X,Y)  |  '+'→ sum
    • '위N군데 전부 후기 총X원' → X
    """
    extra = str(col11) if col11 else ''
    final = str(col12) if col12 else ''

    rev, ref = _parse_review_ref_from_text(extra)
    if rev + ref > 0:
        return rev + ref

    # col11에 없으면 col12에서 fallback
    rev2, ref2 = _parse_review_ref_from_text(final)
    return rev2 + ref2


def _parse_review_ref_from_text(text):
    """텍스트에서 (리뷰금액, 추천금액) 추출"""
    if not text:
        return 0, 0

    # 이벤트 사은품 최대 X (요약 줄) — breakdown 없으면 이 값 사용
    summary_match = re.search(
        r'이벤트\s*사은품\s*(?:최대\s*)?(\d+(?:,\d{3})*|(?:\d+만))\s*원?', text
    )

    # "위 N군데 전부 후기작성시 총 X원"
    total_match = re.search(r'전부.{0,20}총\s*(\d+(?:\.\d+)?)\s*만\s*원?', text)
    if not total_match:
        total_match = re.search(r'전부.{0,20}총\s*(\d{1,3}(?:,\d{3})+)\s*원?', text)

    if total_match:
        v_str = total_match.group(1)
        if '만' in total_match.group(0):
            return int(float(v_str) * 10000), 0
        return int(v_str.replace(',', '')), 0

    # 블로그(X) + 카페(Y) 형태 (additive) — \d+만 먼저 시도
    multi_match = re.search(
        r'(?:블로그|카페)\s*[\(（]?\s*(\d+만|\d+(?:,\d{3})*)\s*원?\s*[\)）]?\s*\+\s*'
        r'(?:블로그|카페)\s*[\(（]?\s*(\d+만|\d+(?:,\d{3})*)\s*원?\s*[\)）]?',
        text, re.IGNORECASE
    )
    if multi_match:
        def to_int(s):
            s = s.strip()
            if s.endswith('만'):
                return int(s[:-1]) * 10000
            return int(s.replace(',', ''))
        return to_int(multi_match.group(1)) + to_int(multi_match.group(2)), 0

    # 리뷰/후기 관련 — look_before=5 (이전 줄 금액 오염 방지)
    review_amounts = amount_near_keyword(
        text, r'후기|리뷰이벤트|리뷰\s*작성|카페\s*후기|블로그\s*후기', 5, 40
    )
    # 카페/블로그 단독 — look_before=5(X만(카페) 형태), look_after=8(같은 줄 금액만)
    # look_after를 짧게 유지해 다음 줄 "총 N만원" 등이 잘못 잡히는 것 방지
    cafe_blog_amounts = amount_near_keyword(
        text, r'카페|블로그', 5, 8
    )
    # 추천 관련 — 키워드 뒤에 금액이 오는 패턴
    referral_amounts = amount_near_keyword(
        text, r'친구\s*추천|지인\s*추천', 0, 30
    )

    review_val = 0
    if review_amounts:
        review_val = max(review_amounts)
    elif cafe_blog_amounts:
        review_val = max(cafe_blog_amounts)

    referral_val = max(referral_amounts) if referral_amounts else 0

    if review_val + referral_val > 0:
        return review_val, referral_val

    # summary 값 fallback
    if summary_match:
        raw = summary_match.group(1).replace(',', '')
        if '만' in raw:
            return int(raw.replace('만', '')) * 10000, 0
        return int(raw), 0

    return 0, 0


# ─────────────────────────────────────────────
# 설치비 — col12 전용
# ─────────────────────────────────────────────
def parse_installation_fee(col12):
    """
    최종 상담 내용(I열)에서 설치비 추출.
    • 금액이 명시된 경우 우선 반환 (설치비지원 여부보다 우선)
    • 금액 없고 '설치비 지원/무상/무료'만 있으면 0 반환
    다양한 구분자 지원: 공백, :, >, /, (, [, 설명어(평일·첫달 등) 최대 20자
    단, '추가설치비·주말설치비·야간설치비'(할증) 앞에 붙은 금액은 무시.
    """
    if not col12:
        return 0
    s = str(col12)

    # ── 1. 명시적 금액 먼저 추출 ──
    # 기본 설치비 키워드 다음에 오는 금액 (설명어 최대 20자 허용)
    # [^\n\d]{0,20} = 개행·숫자 제외 최대 20자 (괄호, 콜론, 슬래시, 한글 설명어 등 허용)
    for pat in [
        # 쉼표 포함 금액 (예: 56,200원, 36,300)
        r'(?<!추가)(?<!주말)(?<!야간)(?:설치비용?|기사출동비|출장비)[^\n\d]{0,20}(\d{1,3}(?:,\d{3})+)(?!\d)',
        # 5-6자리 금액 (예: 56200, 36300)
        r'(?<!추가)(?<!주말)(?<!야간)(?:설치비용?|기사출동비|출장비)[^\n\d]{0,20}(\d{5,6})(?!\d)',
        # 만원 단위 (예: 3.6만원, 5만원)
        r'(?<!추가)(?<!주말)(?<!야간)설치비[^\n\d]{0,15}(\d+(?:\.\d+)?)\s*만\s*원?',
    ]:
        m = re.search(pat, s)
        if m:
            # 숫자 바로 뒤에 '지원/무상/무료'가 오면 업체가 부담 → 스킵
            post = s[m.end():m.end() + 6]
            if re.search(r'지원|무상|무료', post):
                continue
            v_str = m.group(1).replace(',', '')
            try:
                v = float(v_str)
                if '만' in pat and v < 1000:
                    v *= 10000
                v = int(v)
                # 현실적 범위 필터 (설치비는 보통 10,000~200,000원)
                if 10000 <= v <= 200000:
                    return v
            except ValueError:
                pass

    # ── 2. 금액 없을 때만 지원/무상/무료 체크 ──
    if re.search(r'설치비.{0,10}(지원|무상|무료)', s):
        return 0

    return 0


# ─────────────────────────────────────────────
# 파일 처리
# ─────────────────────────────────────────────
def process_file(src_path, dst_path, sheet_name='7차_인터넷',
                 cash_col=9, gift_col=10, extra_col=11, final_col=12,
                 header_row=2, data_start=3):
    """
    K(11) ~ N(14) 열에 분석 결과 삽입:
    K = 추가현금  L = 리뷰보너스  M = 설치비(참고)  N = 총 지원금
    현금·상품권 열(I, J)은 그대로 유지.
    기존 K(그 외 추가 혜택)~N 열은 오른쪽으로 4칸 밀림.
    """
    shutil.copy2(src_path, dst_path)
    wb = openpyxl.load_workbook(dst_path)

    if sheet_name not in wb.sheetnames:
        print(f'  ⚠ 시트 없음: {sheet_name}')
        wb.close()
        return

    ws = wb[sheet_name]

    # 병합 셀 해제
    for merged in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(merged))

    # ── 1단계: 데이터 행 미리 계산 (기존 col11/12 위치 기준) ──
    results = []
    for row_idx in range(data_start, ws.max_row + 1):
        if ws.cell(row_idx, 1).value is None:
            break
        cash  = parse_money_cell(ws.cell(row_idx, cash_col).value)
        gift  = parse_money_cell(ws.cell(row_idx, gift_col).value)
        extra = ws.cell(row_idx, extra_col).value
        final = ws.cell(row_idx, final_col).value

        add_cash = parse_additional_cash(extra)
        rev_ref  = parse_review_referral(extra, final)
        install  = parse_installation_fee(final)
        total    = cash + gift + add_cash + rev_ref  # 설치비 제외

        results.append((row_idx, add_cash, rev_ref, install, total))

    # ── 2단계: K(col11) 위치에 4열 삽입 → K~N 이 새 분석 열 ──
    insert_at = 11  # K열
    ws.insert_cols(insert_at, 4)

    # ── 3단계: 헤더 기록 ──
    headers = [
        ('추가현금',    '추가혜택에서 추가 현금 추출'),
        ('리뷰보너스',  '카페/블로그/네이버 후기 등'),
        ('설치비',      '참고용 — 총지원금 미포함'),
        ('총 지원금',   '현금+상품권+추가현금+리뷰보너스'),
    ]
    for i, (h, _) in enumerate(headers):
        ws.cell(header_row, insert_at + i, h)

    # ── 4단계: 결과 기록 ──
    for row_idx, add_cash, rev_ref, install, total in results:
        ws.cell(row_idx, insert_at,     add_cash if add_cash else None)
        ws.cell(row_idx, insert_at + 1, rev_ref  if rev_ref  else None)
        ws.cell(row_idx, insert_at + 2, install  if install  else None)
        ws.cell(row_idx, insert_at + 3, total)

    wb.save(dst_path)
    wb.close()
    print(f'  완료 — {len(results)}행, K~N열 삽입, 저장: {os.path.basename(dst_path)}')


def process_comparison_sheet(src_path, dst_path,
                              sheets=('인터넷_이지태스크', '인터넷_리얼컨택서비스')):
    """
    지원금 비교 최종 시트(J~M 공란)를 채운다.
      F(6)=현금, G(7)=상품권, H(8)=그 외 추가 혜택, I(9)=최종 상담 내용
      → J(10)=추가현금, K(11)=리뷰+추천 보너스, L(12)=설치비, M(13)=총 지원금
    헤더 행=1, 데이터 시작=2. 열 삽입 없이 기존 공란에 기입.
    """
    shutil.copy2(src_path, dst_path)
    wb = openpyxl.load_workbook(dst_path)

    for sheet_name in sheets:
        if sheet_name not in wb.sheetnames:
            print(f'  ⚠ 시트 없음: {sheet_name}')
            continue

        ws = wb[sheet_name]
        for merged in list(ws.merged_cells.ranges):
            ws.unmerge_cells(str(merged))

        CASH_COL, GIFT_COL = 6, 7   # F, G
        EXTRA_COL, FINAL_COL = 8, 9  # H, I
        J, K, L, M = 10, 11, 12, 13  # 출력 열

        rows_written = 0
        for row_idx in range(2, ws.max_row + 1):
            if ws.cell(row_idx, 1).value is None:
                break

            cash  = parse_money_cell(ws.cell(row_idx, CASH_COL).value)
            gift  = parse_money_cell(ws.cell(row_idx, GIFT_COL).value)
            extra = ws.cell(row_idx, EXTRA_COL).value
            final = ws.cell(row_idx, FINAL_COL).value

            add_cash = parse_additional_cash(extra)
            rev_ref  = parse_review_referral(extra, final)
            install  = parse_installation_fee(final)
            total    = cash + gift + add_cash + rev_ref  # 설치비 제외

            ws.cell(row_idx, J, add_cash)
            ws.cell(row_idx, K, rev_ref)
            ws.cell(row_idx, L, install)
            ws.cell(row_idx, M, total)
            rows_written += 1

        print(f'  [{sheet_name}] {rows_written}행 완료')

    wb.save(dst_path)
    wb.close()
    print(f'  저장: {os.path.basename(dst_path)}')


def process_usim_sheet(wb, sheet_name):
    """
    유심 시트 처리 (인터넷과 동일 로직, 열 위치만 다름)
    I(9)=현금, J(10)=상품권, K(11)=추가혜택, L(12)=최종상담
    → M(13)=추가현금, N(14)=리뷰+추천, O(15)=설치비, P(16)=총지원금
    """
    if sheet_name not in wb.sheetnames:
        print(f'  ⚠ 시트 없음: {sheet_name}')
        return

    ws = wb[sheet_name]
    for merged in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(merged))

    CASH_COL, GIFT_COL = 9, 10    # I, J
    EXTRA_COL, FINAL_COL = 11, 12  # K, L
    M, N, O, P = 13, 14, 15, 16   # 출력 열

    rows_written = 0
    for row_idx in range(2, ws.max_row + 1):
        if ws.cell(row_idx, 1).value is None:
            break

        cash  = parse_money_cell(ws.cell(row_idx, CASH_COL).value)
        gift  = parse_money_cell(ws.cell(row_idx, GIFT_COL).value)
        extra = ws.cell(row_idx, EXTRA_COL).value
        final = ws.cell(row_idx, FINAL_COL).value

        add_cash = parse_additional_cash(extra)
        rev_ref  = parse_review_referral(extra, final)
        install  = parse_installation_fee(final)
        total    = cash + gift + add_cash + rev_ref  # 설치비 제외

        # 기존 셀 덮어쓰기 — None 대신 0으로 써야 기존값 제거됨
        ws.cell(row_idx, M, add_cash)
        ws.cell(row_idx, N, rev_ref)
        ws.cell(row_idx, O, install)
        ws.cell(row_idx, P, total)
        rows_written += 1

    print(f'  [{sheet_name}] {rows_written}행 완료')


def process_appliance_sheet(wb, sheet_name):
    """
    가전 시트 처리
    L(12)=지원금(현금사은품), M(13)=추가혜택(텍스트)
    → N(14)=최종 지원금 = L + 리뷰/추가현금(M에서 파싱)
    반값이벤트(월요금할인) 등은 포함하지 않고, 명시적 현금/리뷰 보너스만 합산.
    """
    if sheet_name not in wb.sheetnames:
        print(f'  ⚠ 시트 없음: {sheet_name}')
        return

    ws = wb[sheet_name]
    for merged in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(merged))

    SUPPORT_COL = 12   # L = 지원금
    EXTRA_COL   = 13   # M = 추가혜택
    N_COL       = 14   # N = 최종 지원금 (출력)

    rows_written = 0
    for row_idx in range(2, ws.max_row + 1):
        if ws.cell(row_idx, 1).value is None:
            break

        support = parse_money_cell(ws.cell(row_idx, SUPPORT_COL).value)
        extra   = ws.cell(row_idx, EXTRA_COL).value

        # M(추가혜택) 텍스트에서 리뷰보너스 + 추가현금만 추출
        # (반값이벤트 등 월요금 할인은 무시)
        rev_bonus = parse_review_referral(extra, None)
        add_cash  = parse_additional_cash(extra)

        total = support + rev_bonus + add_cash

        ws.cell(row_idx, N_COL, total if total else None)
        rows_written += 1

    print(f'  [{sheet_name}] {rows_written}행 완료')


def process_all_sheets(src_path, dst_path):
    """비교 파일의 인터넷·유심·가전 6개 시트 전체 처리"""
    shutil.copy2(src_path, dst_path)
    wb = openpyxl.load_workbook(dst_path)

    print('--- 인터넷 시트 ---')
    for sheet in ('인터넷_이지태스크', '인터넷_리얼컨택서비스'):
        if sheet not in wb.sheetnames:
            print(f'  ⚠ 없음: {sheet}'); continue
        ws = wb[sheet]
        for merged in list(ws.merged_cells.ranges):
            ws.unmerge_cells(str(merged))
        CASH_COL, GIFT_COL = 6, 7
        EXTRA_COL, FINAL_COL = 8, 9
        J, K, L, M = 10, 11, 12, 13
        rows = 0
        for row_idx in range(2, ws.max_row + 1):
            if ws.cell(row_idx, 1).value is None: break
            cash  = parse_money_cell(ws.cell(row_idx, CASH_COL).value)
            gift  = parse_money_cell(ws.cell(row_idx, GIFT_COL).value)
            extra = ws.cell(row_idx, EXTRA_COL).value
            final = ws.cell(row_idx, FINAL_COL).value
            add_cash = parse_additional_cash(extra)
            rev_ref  = parse_review_referral(extra, final)
            install  = parse_installation_fee(final)
            total    = cash + gift + add_cash + rev_ref
            ws.cell(row_idx, J, add_cash)
            ws.cell(row_idx, K, rev_ref)
            ws.cell(row_idx, L, install)
            ws.cell(row_idx, M, total)
            rows += 1
        print(f'  [{sheet}] {rows}행 완료')

    print('--- 유심 시트 ---')
    for sheet in ('유심_이지태스크', '유심_리얼컨택서비스'):
        process_usim_sheet(wb, sheet)

    print('--- 가전 시트 ---')
    for sheet in ('가전_이지태스크', '가전_리얼컨택서비스'):
        process_appliance_sheet(wb, sheet)

    wb.save(dst_path)
    wb.close()
    print(f'\n저장 완료: {os.path.basename(dst_path)}')


if __name__ == '__main__':
    SRC = "/Users/yeongseon/Downloads/인터넷, 가전 상품 조사_지원금 비교_'26년 5월.xlsx"
    DST = "/Users/yeongseon/Downloads/인터넷_지원금비교_완성.xlsx"

    print('=== 전체 시트 처리 (인터넷 + 유심 + 가전) ===')
    process_all_sheets(src_path=SRC, dst_path=DST)
