"""
비교 완성 파일 → 정규화된 DataFrame 변환 모듈
"""
import pandas as pd
import openpyxl
from io import BytesIO


def _safe(v):
    if v is None:
        return 0
    try:
        return int(float(str(v).strip()))
    except Exception:
        return 0


# ── 시트별 컬럼 매핑 ──────────────────────────────────────────
SHEET_SPECS = {
    "인터넷_이지태스크": {
        "source": "이지태스크", "category": "인터넷",
        "telecom": 0, "product": 1, "segment": 2, "vendor": 3,
        "cash": 5, "gift": 6, "add_cash": 9, "review": 10,
        "install": 11, "total": 12,
        "rentre": 13, "diff": 14,
    },
    "인터넷_리얼컨택서비스": {
        "source": "리얼컨택", "category": "인터넷",
        "telecom": 0, "product": 1, "segment": 2, "vendor": 3,
        "cash": 5, "gift": 6, "add_cash": 9, "review": 10,
        "install": 11, "total": 12,
        "rentre": 13, "diff": 14,
    },
    "유심_이지태스크": {
        "source": "이지태스크", "category": "유심",
        "telecom": 0, "product": 1, "segment": 5, "vendor": 6,
        "cash": 8, "gift": 9, "add_cash": 12, "review": 13,
        "install": 14, "total": 15,
        "rentre": 16, "diff": 17,
    },
    "유심_리얼컨택서비스": {
        "source": "리얼컨택", "category": "유심",
        "telecom": 0, "product": 1, "segment": 5, "vendor": 6,
        "cash": 8, "gift": 9, "add_cash": 12, "review": 13,
        "install": 14, "total": 15,
        "rentre": 16, "diff": 17,
    },
    "가전_이지태스크": {
        "source": "이지태스크", "category": "가전",
        "telecom": 1, "product": 2, "segment": 0, "vendor": 4,
        "cash": 11, "gift": None, "add_cash": None, "review": None,
        "install": None, "total": 13,
        "rentre": 15, "diff": 16,
    },
    "가전_리얼컨택서비스": {
        "source": "리얼컨택", "category": "가전",
        "telecom": 1, "product": 2, "segment": 0, "vendor": 4,
        "cash": 11, "gift": None, "add_cash": None, "review": None,
        "install": None, "total": 13,
        "rentre": 15, "diff": 16,
    },
}


def load_comparison_file(file_bytes: bytes):
    """
    비교 완성 파일을 읽어 시트별 DataFrame 딕셔너리 반환.
    각 DataFrame 컬럼: 통신사, 상품명, 구분, 업체명, 현금, 상품권,
                       추가현금, 리뷰보너스, 설치비, 총지원금,
                       렌트리지원금, 격차, source, category
    """
    wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    result = {}

    for sheet_name, spec in SHEET_SPECS.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(min_row=2, max_row=500, values_only=True):
            if row[0] is None:
                break
            def g(idx):
                return _safe(row[idx]) if idx is not None else 0

            total = g(spec["total"])
            if total == 0:
                total = g(spec["cash"]) + g(spec.get("gift") or 0)

            rentre = g(spec["rentre"])
            diff = rentre - total if rentre else None

            rows.append({
                "통신사": str(row[spec["telecom"]] or "").strip(),
                "상품명": str(row[spec["product"]] or "").strip(),
                "구분": str(row[spec["segment"]] or "").strip(),
                "업체명": str(row[spec["vendor"]] or "").strip(),
                "현금": g(spec["cash"]),
                "상품권": g(spec["gift"]) if spec["gift"] is not None else 0,
                "추가현금": g(spec["add_cash"]) if spec["add_cash"] is not None else 0,
                "리뷰보너스": g(spec["review"]) if spec["review"] is not None else 0,
                "설치비": g(spec["install"]) if spec["install"] is not None else 0,
                "총지원금": total,
                "렌트리지원금": rentre,
                "격차": diff,
                "source": spec["source"],
                "category": spec["category"],
            })

        if rows:
            result[sheet_name] = pd.DataFrame(rows)

    wb.close()
    return result
